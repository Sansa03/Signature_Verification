from flask import Flask, render_template, request, jsonify, url_for
import pandas as pd
import torch
import torch.nn as nn
import torchvision.transforms as transforms
from torchvision import models
from PIL import Image
import cv2
import numpy as np
from sklearn.metrics.pairwise import cosine_similarity
import io
from openpyxl import load_workbook
import os

# Set up uploads folder
UPLOAD_FOLDER = os.path.join('static', 'uploads')
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

app = Flask(__name__)

# Load the workbook and specific sheet with signatures
file_path = 'studentsignature.xlsx'  # Path to your Excel file
workbook = load_workbook(file_path)
sheet = workbook.active

# Extract images from the sheet and map them to names
images_data = {}
for image in sheet._images:
    anchor_cell = image.anchor._from.row
    name = sheet.cell(row=anchor_cell + 1, column=1).value  # Assuming names are in the first column
    img_data = io.BytesIO(image._data())
    img = Image.open(img_data)
    images_data[name] = img

# Load pre-trained AlexNet and modify it to output feature embeddings
alexnet = models.alexnet(pretrained=True)
alexnet.classifier = nn.Sequential(*list(alexnet.classifier.children())[:-1])  # Remove last layer to get feature embeddings
alexnet.eval()

# Define transformations for AlexNet input
transform = transforms.Compose([
    transforms.Resize((224, 224)),
    transforms.ToTensor(),
    transforms.Normalize(mean=[0.485, 0.456, 0.406], std=[0.229, 0.224, 0.225])
])

# Preprocess the image to isolate the signature
def preprocess_signature(img):
    img = img.convert('L')  # Convert to grayscale
    img = np.array(img)
    
    # Apply binary threshold to create a black-and-white signature image
    _, binary_img = cv2.threshold(img, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)

    # Remove noise with morphological operations
    kernel = np.ones((3, 3), np.uint8)
    binary_img = cv2.erode(binary_img, kernel, iterations=1)
    binary_img = cv2.dilate(binary_img, kernel, iterations=1)

    # Convert back to PIL format and duplicate channels to create an RGB image
    processed_img = Image.fromarray(binary_img).convert('RGB')
    return processed_img

# Function to extract features using AlexNet
def extract_features(img):
    img = transform(img).unsqueeze(0)  # Transform and add batch dimension
    with torch.no_grad():
        features = alexnet(img)  # Extract features
    return features

# Function to calculate cosine similarity between two feature vectors
def calculate_similarity(features1, features2):
    similarity = cosine_similarity(features1.numpy(), features2.numpy())
    return similarity[0][0] * 100  # Convert to percentage

# Function to save the matched signature image
def save_signature_image(img, filename):
    save_path = os.path.join(UPLOAD_FOLDER, filename)
    img.save(save_path)  # Save the image in the 'uploads' folder
    return filename  # Return the filename for reference in the template

# Main verification function
def verify_signature(name, user_signature_path):
    if name not in images_data:
        return None, "No signature on file."

    # Load and preprocess the stored and user-provided signatures
    stored_signature = preprocess_signature(images_data[name])
    user_signature = preprocess_signature(Image.open(user_signature_path))

    # Extract features
    stored_signature_features = extract_features(stored_signature)
    user_signature_features = extract_features(user_signature)

    # Calculate similarity score
    similarity_score = calculate_similarity(stored_signature_features, user_signature_features)
    if similarity_score >= 85:
        matched_filename = f"{name}_matched.jpeg"
        save_signature_image(images_data[name], matched_filename)  # Save the matched signature
        return similarity_score, matched_filename
    else:
        return similarity_score, None  # No match found

@app.route('/', methods=['GET'])
def index():
    return render_template('index.html')

@app.route('/verify', methods=['POST'])
def verify():
    name = request.form['name']
    uploaded_file = request.files['signature']
    
    if uploaded_file.filename != '':
        saved_filename = os.path.join(UPLOAD_FOLDER, uploaded_file.filename)
        uploaded_file.save(saved_filename)
        user_signature_path = os.path.join('static', 'uploads', uploaded_file.filename)

        # Perform verification
        similarity_score, matched_signature = verify_signature(name, user_signature_path)
        auth_status = ""
        
        if similarity_score is not None and similarity_score >= 85:
            auth_status = f"Success! Match Percentage: {similarity_score:.2f}%"
        else:
            auth_status = "Authentication Fail"
        
        # JSON response with URLs for dynamic table update
        return jsonify({
            "auth_status": auth_status,
            "uploaded_signature_url": url_for('static', filename='uploads/' + uploaded_file.filename),
            "matched_signature_url": url_for('static', filename='uploads/' + matched_signature) if matched_signature else None
        })

    return jsonify({"error": "File upload failed"}), 400

if __name__ == '__main__':
    app.run(debug=True)
